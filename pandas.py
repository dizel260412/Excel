import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl.styles.numbers


# df = pd.read_excel(r"C:\Users\datel3\OneDrive - IKEA\Desktop\Python\pandas\AL010_PG.xlsm", sheet_name='Data',
#                    index_col=None, usecols=['HFB', 'ARTNO', 'EDS'])
# df1 = df[df['ARTNO'] == 30313113]
# df1.to_excel('new.xlsx', index=False)

# df = pd.read_excel(r"C:\Users\datel3\OneDrive - IKEA\Desktop\Python\pandas\AL010_PG.xlsm", sheet_name='Data',
#                    index_col=None, usecols=['HFB', 'ARTNO', 'SALESMETHOD'])
# df1 = df[(df["HFB"] == 1) & (df["SALESMETHOD"] == 1)]
# df1.to_excel('new1.xlsx', index=False)


wb = load_workbook('new1.xlsx')
sheet = wb["Sheet1"]

# print(sheet['d1'].value)  # Значение ячейки


### вывод всех значений столбца В:В с форматированием "00000000"
# for cellObj in sheet['B':'B'][1:]:
#     print(str(cellObj.value).rjust(8, "0"))


# ### вывод всех значений столбца A:A
# for x in range(1, sheet.max_row+1):
#     print(sheet.cell(row=x, column=1).value)


#### вывод всех значений строки 2
#
# for x in range(1, sheet.max_column+1):
#     print(sheet.cell(row=2, column=x).value)


# n = 0
# for row in range(0, sheet.max_row):
#     print(sheet.cell(row + 1, 2).value)  # Вывод заполненых ячеек в 2м ряду (ARTNO)
#     n += 1  # пересчет заполненых ячеек в 2м ряду (ARTNO)
# print(n)

# df = pd.read_excel(r"C:\Users\datel3\OneDrive - IKEA\Desktop\Python\pandas\AL010_PG.xlsm", sheet_name='Data',
#                    index_col=None, usecols=['HFB', 'ARTNO', 'EDS'])


### Создаем новый лист new и записываем в него значения из A:A листа Sheet1
# new_sheet = wb.create_sheet(title='new', index=0)

### Изменяем данные на странице new  в столбцах А:А и В:В
# new_sheet = wb['new']
# i = 1
# for x in range(1, sheet.max_row+1):
#     a = sheet.cell(row=x, column=1).value
#     b = sheet.cell(row=x, column=2).value
#     new_sheet[f'A{i}'] = str(a).rjust(2, "0")
#     new_sheet[f'B{i}'] = str(b).rjust(8, "0")
#     i += 1
# wb.save('new1.xlsx')   ##### Сохраняем изменения в файле
